from django.views import View
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from users.decorators import role_required
from django.utils.decorators import method_decorator
from django.db.models import Sum
from payments.models import Payment
from students.models import Student
import datetime
import io
from datetime import datetime, timedelta
from django.views.generic import ListView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from payments.models import Payment
from groups.models import Group, Enrollment
from students.models import Student
class FinanceReportView2(LoginRequiredMixin, TemplateView):
    template_name = 'finances/finance_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Параметры фильтрации
        period = self.request.GET.get('period', 'month')
        group_id = self.request.GET.get('group')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # Определяем даты периода
        today = timezone.now().date()
        if period == 'day':
            start_date = today
            end_date = today
        elif period == 'week':
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
        elif period == 'month':
            start_date = today.replace(day=1)
            end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Если указаны конкретные даты
        if self.request.GET.get('start_date') and self.request.GET.get('end_date'):
            start_date = datetime.strptime(self.request.GET['start_date'], '%Y-%m-%d').date()
            end_date = datetime.strptime(self.request.GET['end_date'], '%Y-%m-%d').date()
        
        # Получаем платежи за период
        payments = Payment.objects.filter(
            date__gte=start_date,
            date__lte=end_date
        )
        
        # Фильтрация по группе
        if group_id:
            payments = payments.filter(group_id=group_id)
        
        # Агрегации
        total_income = payments.aggregate(Sum('amount'))['amount__sum'] or 0
        payments_count = payments.count()
        
        # Доход по группам
        income_by_group = payments.values('group__name').annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('-total')
        
        # Студенты с долгами - ИСПРАВЛЕННЫЙ МЕТОД
        debt_students = self.get_debt_students()
        
        # Расчет количества дней в периоде для отображения
        period_days = (end_date - start_date).days + 1
        
        context.update({
            'payments': payments.select_related('student', 'group'),
            'total_income': total_income,
            'payments_count': payments_count,
            'income_by_group': income_by_group,
            'debt_students': debt_students,
            'groups': Group.objects.all(),
            'period': period,
            'start_date': start_date,
            'end_date': end_date,
            'selected_group': group_id,
            'period_days': period_days,  # Добавляем для отображения
        })
        
        return context
    
    def get_debt_students(self):
        """Возвращает студентов с задолженностями - ИСПРАВЛЕННАЯ ВЕРСИЯ"""
        debt_students = []
        
        # Получаем всех активных студентов
        active_students = Student.objects.filter(status='active')
        
        for student in active_students:
            # Получаем все зачисления студента
            enrollments = Enrollment.objects.filter(student=student).select_related('group')
            
            for enrollment in enrollments:
                # Получаем текущий персональный месяц студента
                current_month = enrollment.get_personal_month()
                
                # Проверяем оплату за текущий месяц
                payment_status = Payment.get_student_payment_status(
                    student,
                    enrollment.group,
                    current_month
                )
                
                # ОТЛАДОЧНАЯ ИНФОРМАЦИЯ
                print(f"DEBUG: Студент {student.full_name}, Группа {enrollment.group.name}")
                print(f"DEBUG: Текущий месяц: {current_month}")
                print(f"DEBUG: Оплачено: {payment_status['total_paid']}, Требуется: {payment_status['monthly_price']}")
                print(f"DEBUG: Полностью оплачено: {payment_status['is_fully_paid']}")
                
                # Если студент не оплатил полностью текущий месяц И у него есть прогресс
                if not payment_status['is_fully_paid'] and enrollment.lessons_attended > 0:
                    debt_students.append({
                        'student': student,
                        'group': enrollment.group,
                        'debt_amount': payment_status['remaining_amount'],
                        'paid_amount': payment_status['total_paid'],
                        'required_amount': payment_status['monthly_price'],
                        'current_month': current_month,
                        'lessons_attended': enrollment.lessons_attended
                    })
                    print(f"DEBUG: Добавлен в список должников")
        
        print(f"DEBUG: Всего должников: {len(debt_students)}")
        return debt_students
    
# class FinanceReportView2(LoginRequiredMixin, TemplateView):
#     template_name = 'finances/finance_report.html'
    
#     def get_context_data(self, **kwargs):
#         context = super().get_context_data(**kwargs)
        
#         # Параметры фильтрации
#         period = self.request.GET.get('period', 'month')
#         group_id = self.request.GET.get('group')
#         start_date = self.request.GET.get('start_date')
#         end_date = self.request.GET.get('end_date')
        
#         # Определяем даты периода
#         today = timezone.now().date()
#         if period == 'day':
#             start_date = today
#             end_date = today
#         elif period == 'week':
#             start_date = today - timedelta(days=today.weekday())
#             end_date = start_date + timedelta(days=6)
#         elif period == 'month':
#             start_date = today.replace(day=1)
#             end_date = (start_date + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
#         # Если указаны конкретные даты
#         if self.request.GET.get('start_date') and self.request.GET.get('end_date'):
#             start_date = datetime.strptime(self.request.GET['start_date'], '%Y-%m-%d').date()
#             end_date = datetime.strptime(self.request.GET['end_date'], '%Y-%m-%d').date()
        
#         # Получаем платежи за период
#         payments = Payment.objects.filter(
#             date__gte=start_date,
#             date__lte=end_date
#         )
        
#         # Фильтрация по группе
#         if group_id:
#             payments = payments.filter(group_id=group_id)
        
#         # Агрегации
#         total_income = payments.aggregate(Sum('amount'))['amount__sum'] or 0
#         payments_count = payments.count()
        
#         # Доход по группам
#         income_by_group = payments.values('group__name').annotate(
#             total=Sum('amount'),
#             count=Count('id')
#         ).order_by('-total')
        
#         # Студенты с долгами
#         debt_students = self.get_debt_students()
        
#         context.update({
#             'payments': payments.select_related('student', 'group'),
#             'total_income': total_income,
#             'payments_count': payments_count,
#             'income_by_group': income_by_group,
#             'debt_students': debt_students,
#             'groups': Group.objects.all(),
#             'period': period,
#             'start_date': start_date,
#             'end_date': end_date,
#             'selected_group': group_id,
#         })
        
#         return context
    
#     def get_debt_students(self):
#         """Возвращает студентов с задолженностями"""
#         debt_students = []
#         enrollments = Enrollment.objects.select_related('student', 'group')
        
#         for enrollment in enrollments:
#             next_month = enrollment.get_next_personal_month()
#             payment_status = Payment.get_student_payment_status(
#                 enrollment.student,
#                 enrollment.group,
#                 next_month
#             )
            
#             if not payment_status['is_fully_paid'] and enrollment.should_check_personal_payment():
#                 debt_students.append({
#                     'student': enrollment.student,
#                     'group': enrollment.group,
#                     'debt_amount': payment_status['remaining_amount'],
#                     'paid_amount': payment_status['total_paid'],
#                     'required_amount': payment_status['monthly_price']
#                 })
        
#         return debt_students

class ExportFinanceReportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # Получаем параметры фильтрации
        period = request.GET.get('period', 'month')
        group_id = request.GET.get('group')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # Создаем workbook
        wb = Workbook()
        
        # Лист с общим отчетом
        ws1 = wb.active
        ws1.title = "Финансовый отчет"
        
        # Заголовки
        headers = ['Дата', 'Студент', 'Группа', 'Сумма', 'Месяц оплаты', 'Примечания']
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Получаем данные
        payments = Payment.objects.all()
        if start_date and end_date:
            payments = payments.filter(date__gte=start_date, date__lte=end_date)
        if group_id:
            payments = payments.filter(group_id=group_id)
        
        # Заполняем данными
        for row, payment in enumerate(payments, 2):
            ws1.cell(row=row, column=1, value=payment.date.strftime('%d.%m.%Y'))
            ws1.cell(row=row, column=2, value=payment.student.full_name)
            ws1.cell(row=row, column=3, value=payment.group.name)
            ws1.cell(row=row, column=4, value=float(payment.amount))
            ws1.cell(row=row, column=5, value=payment.payment_month_number)
            ws1.cell(row=row, column=6, value=payment.notes or '')
        
        # Авто-ширина колонок
        for column in ws1.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # Лист с задолженностями
        ws2 = wb.create_sheet("Задолженности")
        
        debt_headers = ['Студент', 'Группа', 'Задолженность', 'Оплачено', 'Требуется']
        for col, header in enumerate(debt_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFDDDD", end_color="FFDDDD", fill_type="solid")
        
        # Получаем студентов с долгами
        debt_students = self.get_debt_students()
        for row, debt in enumerate(debt_students, 2):
            ws2.cell(row=row, column=1, value=debt['student'].full_name)
            ws2.cell(row=row, column=2, value=debt['group'].name)
            ws2.cell(row=row, column=3, value=float(debt['debt_amount']))
            ws2.cell(row=row, column=4, value=float(debt['paid_amount']))
            ws2.cell(row=row, column=5, value=float(debt['required_amount']))
        
        # Авто-ширина для второго листа
        for column in ws2.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws2.column_dimensions[column_letter].width = adjusted_width
        
        # Создаем response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"financial_report_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        # Сохраняем workbook в response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.getvalue())
        
        return response
    
    def get_debt_students(self):
        """Возвращает студентов с задолженностями"""
        debt_students = []
        enrollments = Enrollment.objects.select_related('student', 'group')
        
        for enrollment in enrollments:
            next_month = enrollment.get_next_personal_month()
            payment_status = Payment.get_student_payment_status(
                enrollment.student,
                enrollment.group,
                next_month
            )
            
            if not payment_status['is_fully_paid'] and enrollment.should_check_personal_payment():
                debt_students.append({
                    'student': enrollment.student,
                    'group': enrollment.group,
                    'debt_amount': payment_status['remaining_amount'],
                    'paid_amount': payment_status['total_paid'],
                    'required_amount': payment_status['monthly_price']
                })
        
        return debt_students
@method_decorator(role_required(['admin', 'accountant']), name='dispatch')
class FinanceReportView(LoginRequiredMixin, TemplateView):
    template_name = 'finances/report.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Параметры фильтрации
        period = self.request.GET.get('period', 'month')
        student_id = self.request.GET.get('student')
        group_id = self.request.GET.get('group')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        # Базовый queryset
        payments = Payment.objects.all()

        # Применяем фильтры
        if student_id:
            payments = payments.filter(student_id=student_id)
        if group_id:
            payments = payments.filter(group_id=group_id)
        if start_date and end_date:
            payments = payments.filter(date__range=[start_date, end_date])
        else:
            # Фильтр по периоду
            today = datetime.date.today()
            if period == 'day':
                payments = payments.filter(date=today)
            elif period == 'week':
                start_week = today - datetime.timedelta(days=today.weekday())
                end_week = start_week + datetime.timedelta(days=6)
                payments = payments.filter(date__range=[start_week, end_week])
            else:  # month
                payments = payments.filter(date__year=today.year, date__month=today.month)

        # Расчеты
        total_income = payments.aggregate(Sum('amount'))['amount__sum'] or 0
        debt_students = Student.objects.filter(status='debt')

        context.update({
            'payments': payments,
            'total_income': total_income,
            'debt_students': debt_students,
            'period': period,
        })
        return context